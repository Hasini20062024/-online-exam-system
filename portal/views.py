import csv
import io
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import load_workbook
import xlrd

from .forms import CourseForm, ExamForm, FacultyCreationForm, LoginForm, QuestionForm, StudentRegistrationForm
from .models import (
    Course,
    Enrollment,
    Exam,
    ExamAttempt,
    ExamRanking,
    Profile,
    Question,
    build_attempt_answers,
    recompute_exam_rankings,
)


QUESTION_UPLOAD_HEADERS = ['question_text', 'option_a', 'option_b', 'option_c', 'option_d', 'correct_answer', 'difficulty']
QUESTION_UPLOAD_FIELDS = ['question_text', 'option_a', 'option_b', 'option_c', 'option_d', 'correct_answer', 'difficulty']


def normalize_excel_value(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def read_question_rows(upload):
    name = upload.name.lower()
    if name.endswith('.xlsx'):
        workbook = load_workbook(filename=io.BytesIO(upload.read()), data_only=True)
        sheet = workbook.active
        return [[normalize_excel_value(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    if name.endswith('.xls'):
        workbook = xlrd.open_workbook(file_contents=upload.read())
        sheet = workbook.sheet_by_index(0)
        rows = []
        for row_index in range(sheet.nrows):
            rows.append([normalize_excel_value(sheet.cell_value(row_index, col_index)) for col_index in range(sheet.ncols)])
        return rows
    raise ValueError('Please upload an Excel file in .xlsx or .xls format.')


def validate_question_excel_rows(rows):
    if not rows:
        return None, ['The uploaded file is empty.']

    header = [normalize_excel_value(value).lower() for value in rows[0][:7]]
    if header != QUESTION_UPLOAD_HEADERS:
        return None, ['Row 1 must contain: question_text, option_a, option_b, option_c, option_d, correct_answer, difficulty.']

    cleaned_rows = []
    errors = []
    for row_number, row in enumerate(rows[1:], start=2):
        values = [normalize_excel_value(row[index]) if index < len(row) else '' for index in range(7)]
        if not any(values):
            continue

        row_data = dict(zip(QUESTION_UPLOAD_FIELDS, values))
        row_errors = []
        if any(not row_data[field] for field in QUESTION_UPLOAD_FIELDS):
            row_errors.append('all fields must be non-empty')
        if row_data['correct_answer'] not in {'A', 'B', 'C', 'D'}:
            row_errors.append('correct_answer must be A/B/C/D')
        if row_data['difficulty'] not in {Question.DIFFICULTY_EASY, Question.DIFFICULTY_MEDIUM, Question.DIFFICULTY_HARD}:
            row_errors.append('difficulty must be Easy/Medium/Hard')

        if row_errors:
            errors.append(f'Row {row_number}: ' + '; '.join(row_errors))
        else:
            cleaned_rows.append(row_data)

    if not cleaned_rows and not errors:
        errors.append('The file has no question rows after the header.')
    return cleaned_rows, errors


def home(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return redirect('login')


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    form = LoginForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        login(request, form.cleaned_data['user'])
        return redirect('dashboard')
    return render(request, 'portal/login.html', {'form': form})


def logout_view(request):
    logout(request)
    return redirect('login')


def register_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    form = StudentRegistrationForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = form.save()
        login(request, user)
        messages.success(request, 'Student registration successful.')
        return redirect('dashboard')
    return render(request, 'portal/register.html', {'form': form})


def get_role(user):
    if user.is_superuser:
        return Profile.ROLE_ADMIN
    return getattr(user.profile, 'role', Profile.ROLE_STUDENT)


def build_student_context(user):
    profile = user.profile
    enrolled_course_ids = user.enrollments.values_list('course_id', flat=True)
    available_courses = Course.objects.exclude(id__in=enrolled_course_ids)
    if profile.academic_year:
        available_courses = available_courses.filter(academic_year=profile.academic_year)

    my_courses = Course.objects.filter(enrollments__student=user).prefetch_related('faculty_members')
    my_exams = Exam.objects.filter(course__enrollments__student=user).select_related('course').distinct()
    attempts = ExamAttempt.objects.filter(student=user).select_related('exam', 'exam__course')
    attempt_map = {attempt.exam_id: attempt for attempt in attempts}
    ranking_map = {ranking.exam_id: ranking for ranking in ExamRanking.objects.filter(student=user)}

    exam_rows = []
    for exam in my_exams:
        attempt = attempt_map.get(exam.id)
        status = exam.status()
        exam_rows.append(
            {
                'exam': exam,
                'status': status,
                'attempt': attempt,
                'ranking': ranking_map.get(exam.id),
                'can_attempt': status == 'Ongoing' and not (attempt and attempt.status == ExamAttempt.STATUS_SUBMITTED),
            }
        )

    return {
        'available_courses': available_courses,
        'my_courses': my_courses,
        'exam_rows': exam_rows,
        'completed_attempts': attempts.filter(status=ExamAttempt.STATUS_SUBMITTED),
    }


def build_faculty_context(user):
    assigned_courses = user.assigned_courses.all().prefetch_related('enrollments')
    faculty_course_rows = []
    for course in assigned_courses:
        faculty_course_rows.append(
            {
                'course': course,
                'enrolled_count': course.enrolled_count,
                'exam_count': course.exams.filter(faculty=user).count(),
            }
        )
    faculty_exam_rows = []
    for exam in Exam.objects.filter(faculty=user).select_related('course'):
        faculty_exam_rows.append(
            {
                'exam': exam,
                'status': exam.status(),
            }
        )
    return {
        'assigned_courses': assigned_courses,
        'faculty_course_rows': faculty_course_rows,
        'question_form': QuestionForm(faculty_user=user),
        'exam_form': ExamForm(faculty_user=user),
        'questions': Question.objects.filter(created_by=user).select_related('course'),
        'exams': faculty_exam_rows,
        'upload_years': assigned_courses.values_list('academic_year', flat=True).distinct(),
        'upload_summary': None,
    }


def build_admin_context():
    faculty_profiles = Profile.objects.filter(role=Profile.ROLE_FACULTY).select_related('user')
    courses = Course.objects.prefetch_related('faculty_members', 'enrollments')
    course_rows = []
    for course in courses:
        assigned_ids = set(course.faculty_members.values_list('id', flat=True))
        available_faculty = faculty_profiles.exclude(user_id__in=assigned_ids)
        course_rows.append({'course': course, 'available_faculty': available_faculty})

    return {
        'course_form': CourseForm(),
        'faculty_form': FacultyCreationForm(),
        'courses': courses,
        'course_rows': course_rows,
        'faculty_profiles': faculty_profiles,
        'new_faculty_credentials': None,
        'summary': {
            'total_courses': Course.objects.count(),
            'total_faculty': Profile.objects.filter(role=Profile.ROLE_FACULTY).count(),
            'total_students': Profile.objects.filter(role=Profile.ROLE_STUDENT).count(),
            'total_exams': Exam.objects.count(),
        },
    }


@login_required
def dashboard_view(request):
    role = get_role(request.user)
    context = {'role': role}
    if role == Profile.ROLE_ADMIN:
        context.update(build_admin_context())
        context['new_faculty_credentials'] = request.session.pop('new_faculty_credentials', None)
    elif role == Profile.ROLE_FACULTY:
        context.update(build_faculty_context(request.user))
        context['upload_summary'] = request.session.pop('upload_summary', None)
    else:
        context.update(build_student_context(request.user))
    return render(request, 'portal/dashboard.html', context)


@login_required
def enroll_in_course(request, course_id):
    if request.method != 'POST' or get_role(request.user) != Profile.ROLE_STUDENT:
        raise Http404
    course = get_object_or_404(Course, id=course_id)
    Enrollment.objects.get_or_create(course=course, student=request.user)
    messages.success(request, f'Enrolled in {course.code}.')
    return redirect('dashboard')


@login_required
def create_faculty(request):
    if request.method != 'POST' or get_role(request.user) != Profile.ROLE_ADMIN:
        raise Http404
    form = FacultyCreationForm(request.POST)
    if form.is_valid():
        user, temp_password = form.save()
        request.session['new_faculty_credentials'] = {
            'full_name': user.profile.full_name,
            'username': user.username,
            'password': temp_password,
        }
        messages.success(request, f'Faculty created for {user.profile.full_name}.')
    else:
        messages.error(request, form.errors.as_text())
    return redirect('dashboard')


@login_required
def create_course(request):
    if request.method != 'POST' or get_role(request.user) != Profile.ROLE_ADMIN:
        raise Http404
    form = CourseForm(request.POST)
    if form.is_valid():
        form.save()
        messages.success(request, 'Course created successfully.')
    else:
        messages.error(request, form.errors.as_text())
    return redirect('dashboard')


@login_required
def assign_faculty_to_course(request, course_id):
    if request.method != 'POST' or get_role(request.user) != Profile.ROLE_ADMIN:
        raise Http404

    course = get_object_or_404(Course, id=course_id)
    faculty_user_id = request.POST.get('faculty_user_id')
    if not faculty_user_id or not faculty_user_id.isdigit():
        messages.error(request, 'Please select a faculty member to assign.')
        return redirect('dashboard')

    faculty = get_object_or_404(User, id=int(faculty_user_id), profile__role=Profile.ROLE_FACULTY)
    course.faculty_members.add(faculty)
    messages.success(request, f'{faculty.profile.full_name} assigned to {course.code}.')
    return redirect('dashboard')


@login_required
def create_question(request):
    if request.method != 'POST' or get_role(request.user) != Profile.ROLE_FACULTY:
        raise Http404
    form = QuestionForm(request.POST, faculty_user=request.user)
    if form.is_valid():
        question = form.save(commit=False)
        question.created_by = request.user
        question.save()
        messages.success(request, 'Question uploaded to the bank.')
    else:
        messages.error(request, form.errors.as_text())
    return redirect('dashboard')


@login_required
def upload_questions_excel(request):
    if request.method != 'POST' or get_role(request.user) != Profile.ROLE_FACULTY:
        raise Http404

    academic_year = (request.POST.get('academic_year') or '').strip()
    course_id = request.POST.get('course_id')
    upload = request.FILES.get('question_file')

    if not academic_year or not course_id or not upload:
        messages.error(request, 'Academic year, course, and Excel file are all required.')
        return redirect('dashboard')

    course = get_object_or_404(Course, id=course_id, faculty_members=request.user)
    if course.academic_year != academic_year:
        messages.error(request, 'Selected course does not belong to the chosen academic year.')
        return redirect('dashboard')

    try:
        rows = read_question_rows(upload)
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect('dashboard')

    cleaned_rows, errors = validate_question_excel_rows(rows)
    if errors:
        messages.error(request, 'Upload rejected: ' + ' | '.join(errors))
        return redirect('dashboard')

    created_questions = []
    counts = {
        Question.DIFFICULTY_EASY: 0,
        Question.DIFFICULTY_MEDIUM: 0,
        Question.DIFFICULTY_HARD: 0,
    }
    for row_data in cleaned_rows:
        created_questions.append(
            Question(
                course=course,
                created_by=request.user,
                question_text=row_data['question_text'],
                option_a=row_data['option_a'],
                option_b=row_data['option_b'],
                option_c=row_data['option_c'],
                option_d=row_data['option_d'],
                correct_answer=row_data['correct_answer'],
                difficulty=row_data['difficulty'],
            )
        )
        counts[row_data['difficulty']] += 1

    Question.objects.bulk_create(created_questions)
    request.session['upload_summary'] = {
        'total': len(created_questions),
        'easy': counts[Question.DIFFICULTY_EASY],
        'medium': counts[Question.DIFFICULTY_MEDIUM],
        'hard': counts[Question.DIFFICULTY_HARD],
        'course': course.code,
    }
    messages.success(request, f'Upload successful. {len(created_questions)} questions added to {course.code}.')
    return redirect('dashboard')


@login_required
def create_exam(request):
    if request.method != 'POST' or get_role(request.user) != Profile.ROLE_FACULTY:
        raise Http404
    form = ExamForm(request.POST, faculty_user=request.user)
    if form.is_valid():
        exam = form.save(commit=False)
        exam.faculty = request.user
        exam.save()
        messages.success(request, 'Exam scheduled successfully.')
    else:
        messages.error(request, form.errors.as_text())
    return redirect('dashboard')


@login_required
def start_exam(request, exam_id):
    if get_role(request.user) != Profile.ROLE_STUDENT:
        raise Http404
    exam = get_object_or_404(Exam.objects.select_related('course'), id=exam_id)
    if not Enrollment.objects.filter(course=exam.course, student=request.user).exists():
        messages.error(request, 'You are not enrolled in this course.')
        return redirect('dashboard')
    if exam.status() != 'Ongoing':
        messages.error(request, 'Exam is not ongoing.')
        return redirect('dashboard')

    attempt, created = ExamAttempt.objects.get_or_create(exam=exam, student=request.user)
    if attempt.status == ExamAttempt.STATUS_SUBMITTED:
        return redirect('result_detail', attempt_id=attempt.id)
    if created:
        build_attempt_answers(attempt)

    return render(
        request,
        'portal/exam_take.html',
        {
            'exam': exam,
            'attempt': attempt,
            'answers': attempt.answers.select_related('question'),
            'remaining_seconds': attempt.remaining_seconds,
        },
    )


@login_required
def submit_exam(request, attempt_id):
    if request.method != 'POST' or get_role(request.user) != Profile.ROLE_STUDENT:
        raise Http404
    attempt = get_object_or_404(ExamAttempt, id=attempt_id, student=request.user)
    if attempt.status == ExamAttempt.STATUS_SUBMITTED:
        return redirect('result_detail', attempt_id=attempt.id)

    score = 0
    answers = attempt.answers.select_related('question')
    for answer in answers:
        value = request.POST.get(f'question_{answer.id}')
        answer.selected_option = int(value) if value and value.isdigit() else None
        answer.is_correct = answer.selected_option == answer.correct_option
        if answer.is_correct:
            score += 1
        answer.save(update_fields=['selected_option', 'is_correct'])

    total_questions = answers.count() or 1
    submitted_at = timezone.localtime()
    attempt.status = ExamAttempt.STATUS_SUBMITTED
    attempt.submitted_at = submitted_at
    attempt.score = score
    attempt.percentage = Decimal(score * 100 / total_questions).quantize(Decimal('0.01'))
    attempt.time_taken = submitted_at - attempt.started_at
    attempt.save(update_fields=['status', 'submitted_at', 'score', 'percentage', 'time_taken'])
    recompute_exam_rankings(attempt.exam)
    messages.success(request, 'Exam submitted.')
    return redirect('result_detail', attempt_id=attempt.id)


@login_required
def result_detail(request, attempt_id):
    attempt = get_object_or_404(
        ExamAttempt.objects.select_related('exam', 'exam__course', 'student', 'student__profile'),
        id=attempt_id,
    )
    role = get_role(request.user)
    if role == Profile.ROLE_STUDENT and attempt.student != request.user:
        raise Http404
    if role == Profile.ROLE_FACULTY and attempt.exam.faculty != request.user:
        raise Http404
    ranking = ExamRanking.objects.filter(exam=attempt.exam, student=attempt.student).first()
    return render(
        request,
        'portal/result_detail.html',
        {
            'attempt': attempt,
            'ranking': ranking,
            'answers': attempt.answers.select_related('question'),
        },
    )


@login_required
def faculty_exam_results(request, exam_id):
    if get_role(request.user) != Profile.ROLE_FACULTY:
        raise Http404
    exam = get_object_or_404(Exam.objects.select_related('course'), id=exam_id, faculty=request.user)
    rankings = ExamRanking.objects.filter(exam=exam).select_related('student', 'student__profile')
    return render(request, 'portal/faculty_results.html', {'exam': exam, 'rankings': rankings})


@login_required
def export_exam_results_csv(request, exam_id):
    if get_role(request.user) != Profile.ROLE_FACULTY:
        raise Http404
    exam = get_object_or_404(Exam, id=exam_id, faculty=request.user)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{exam.title}_results.csv"'
    writer = csv.writer(response)
    writer.writerow(['Rank', 'Student Full Name', 'Username', 'Score', 'Percentage', 'Submitted At'])
    for ranking in ExamRanking.objects.filter(exam=exam).select_related('student', 'student__profile'):
        writer.writerow(
            [
                ranking.rank,
                ranking.student.profile.full_name,
                ranking.student.username,
                ranking.score,
                ranking.percentage,
                timezone.localtime(ranking.submitted_at).strftime('%Y-%m-%d %H:%M:%S'),
            ]
        )
    return response
